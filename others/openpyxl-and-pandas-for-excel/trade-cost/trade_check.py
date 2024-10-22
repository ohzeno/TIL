from openpyxl import load_workbook
import time
import pandas as pd

t1 = time.time()
load_wb = load_workbook("trade_list.xlsx")  # 엑셀 로드
ws = load_wb[load_wb.sheetnames[0]]  # 첫 시트 가져오기
years = {}  # 정보 담을 딕셔너리
# 연도 따라 딕셔너리 추가
for i in range(2019, 2023):
    years[str(i)] = {
        '코스피': {
            '매수': {
                '수수료': [],
                '거래세': [],
                '세금': []
            },
            '매도': {
                '수수료': [],
                '거래세': [],
                '세금': []
            }
        },
        '코스닥': {
            '매수': {
                '수수료': [],
                '거래세': [],
                '세금': []
            },
            '매도': {
                '수수료': [],
                '거래세': [],
                '세금': []
            }
        }
    }
tmp = []
for row in ws.rows:  # 시트 한 행씩 가져오기
    # 첫 두 줄은 헤더임. 무시
    if row[1].row == 1 or row[1].row == 2:
        continue
    # 거래일, 거래종류, 종목명은 2줄씩 병합되어 있으므로
    # 두 줄을 모아서 체크하기 위해 tmp에 넣는다.
    if row[1].row % 2 == 1:
        tmp.append(row)
    else:
        tmp.append(row)
        # 2줄 모은 후 수수료, 거래세, 세금칸 체크 후
        # 비용이 있다면 분류해서 비용/거래금액을 추가해준다.
        if int(tmp[0][5].internal_value.replace(',', '')):
            year = tmp[0][0].internal_value[:4]
            bs = tmp[0][1].internal_value[5:7]
            if tmp[0][1].internal_value[:3] == '거래소':
                ty = '코스피'
            elif tmp[0][1].internal_value[:3] == '코스닥':
                ty = '코스닥'
            tmp_val = int(tmp[0][5].internal_value.replace(',', ''))/int(tmp[0][4].internal_value.replace(',', ''))
            years[year][ty][bs]['수수료'].append(tmp_val)
        if int(tmp[1][5].internal_value.replace(',', '')):
            year = tmp[0][0].internal_value[:4]
            bs = tmp[0][1].internal_value[5:7]
            if tmp[0][1].internal_value[:3] == '거래소':
                ty = '코스피'
            elif tmp[0][1].internal_value[:3] == '코스닥':
                ty = '코스닥'
            tmp_val = int(tmp[1][5].internal_value.replace(',', '')) / int(tmp[0][4].internal_value.replace(',', ''))
            years[year][ty][bs]['거래세'].append(tmp_val)
        if int(tmp[1][6].internal_value.replace(',', '')):
            year = tmp[0][0].internal_value[:4]
            bs = tmp[0][1].internal_value[5:7]
            if tmp[0][1].internal_value[:3] == '거래소':
                ty = '코스피'
            elif tmp[0][1].internal_value[:3] == '코스닥':
                ty = '코스닥'
            tmp_val = int(tmp[1][6].internal_value.replace(',', '')) / int(tmp[0][4].internal_value.replace(',', ''))
            years[year][ty][bs]['세금'].append(tmp_val)
        # 두 줄 처리했으니 tmp를 비워준다.
        tmp = []

with pd.ExcelWriter('test.xlsx') as writer:
    # test.xlsx파일을 만들고 데이터를 연도별 코스피, 코스닥으로 나눠 각각 시트에 저장한다.
    for i in range(2019, 2023):
        pd.DataFrame(years[f'{str(i)}']['코스피']).transpose().to_excel(writer, sheet_name=f'{str(i)}_코스피')
        pd.DataFrame(years[f'{str(i)}']['코스닥']).transpose().to_excel(writer, sheet_name=f'{str(i)}_코스닥')
t2 = time.time()
print(f'소요시간: {t2-t1}')
